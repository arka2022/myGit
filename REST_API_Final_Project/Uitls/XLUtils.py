import openpyxl
loc="/Users/arkapdas/PycharmProjects/REST_API_Final_Project/TestData/UserDatabase.xlsx"

class XLRead:
    def UserData_readXL(loc,row,col):
        wb = openpyxl.load_workbook(loc)
        sh1 = wb['UserData']
        mr = sh1.max_row  # maximum row in sheet
        print(mr)
        mc = sh1.max_column  # max column in sheet
        print(mc)
        return sh1.cell(row,col).value

    def UserData_writeXL(loc,col,value):
        wb = openpyxl.load_workbook(loc)
        sh1 = wb['UserData']
        mr = sh1.max_row  # maximum row in sheet
        print(mr)
        mc = sh1.max_column  # max column in sheet
        print(mc)
        sh1.cell(row=mr + 1, column=col, value=value)
        wb.save("UserDatabase.xlsx")






