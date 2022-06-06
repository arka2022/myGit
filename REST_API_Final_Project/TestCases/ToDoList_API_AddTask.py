import  requests
import json
import jsonpath
import openpyxl

from TestCases.ToDoList_API_UserUpdate import TestUser
from Uitls.XLUtils import XLRead
from Uitls.readText import Read_Write_Txt
from Uitls.readUrl import URLS
import logging


logger = logging.getLogger()
fhandler = logging.FileHandler(filename='/Users/arkapdas/PycharmProjects/REST_API_Final_Project/Log/Addtask.txt', mode='a+')
formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
fhandler.setFormatter(formatter)
logger.addHandler(fhandler)
logger.setLevel(logging.DEBUG)

class TestAPI_addTask:

    def test_login(self):
        url=URLS.login_User_url
        file = open("/Users/arkapdas/PycharmProjects/REST_API_Final_Project/TestData/login_user.json")
        json_ip = file.read()
        request_json = json.loads(json_ip)
        response = requests.post(url, json=request_json)
        assert response.status_code==200
        logger.info('Status code:  %s'%(response.status_code))

        respo_json=json.loads(response.text)
        t=jsonpath.jsonpath(respo_json,'token')
        Read_Write_Txt.token_write(self,str(t[0]))



    def test_addtask(self):
        #TestUser.test_login(self)
        url = URLS.add_task_url
        # Load/read json file
        file = open("/Users/arkapdas/PycharmProjects/REST_API_Final_Project/TestData/add_task.json")
        json_ip = file.read()
        request_json = json.loads(json_ip)
        # Set header
        tok = Read_Write_Txt.token_read(self)
        headers_dict = {"Authorization": "Bearer " + tok}

        resp = requests.post(url, json=request_json, headers=headers_dict)
        assert resp.status_code == 201
        logger.info('Status code:  %s' % (resp.status_code))

    def test_addmultipleTasks(self):
        url = URLS.add_task_url
        loc = "/Users/arkapdas/PycharmProjects/REST_API_Final_Project/TestData/UserDatabase.xlsx"
        wb = openpyxl.load_workbook(loc)
        sh1 = wb['Tasks']
        mr = sh1.max_row  # maximum row in sheet
        mc = sh1.max_column  # max column in sheet

        tok = Read_Write_Txt.token_read(self)
        headers_dict = {"Authorization": "Bearer " + tok}


        for i in range(1,mr):
            json_ip={sh1.cell(i,1).value: sh1.cell(i,2).value}
            print(json_ip)
            resp = requests.post(url, json=json_ip, headers=headers_dict)



    def test_getAllTask(self):
        url=URLS.getAllTask_url
        tok = Read_Write_Txt.token_read(self)
        headers_dict = {"Authorization": "Bearer " + tok}
        resp = requests.get(url,headers=headers_dict)
        assert resp.status_code == 200
        logger.info('Status code:  %s' % (resp.status_code))

    def test_pagination(self):
        url = URLS.pagination_url
        tok = Read_Write_Txt.token_read(self)
        headers_dict = {"Authorization": "Bearer " + tok}
        # https://api-nodejs-todolist.herokuapp.com/task?limit=2&skip=10
        limit=str(2)
        skip=str(1)
        url2=url+url+'limit='+limit+'&skip='+skip
        print(url2)
        resp = requests.get(url2, headers=headers_dict)
        print(resp.text)
        assert resp.status_code == 200
        logger.info('Status code:  %s' % (resp.status_code))















