import sys
import os
import openpyxl

loc=("/Users/arkapdas/PycharmProjects/projectFlipkart/TestData/FoodSwipeDatabase.xlsx")
wb=openpyxl.load_workbook("/Users/arkapdas/PycharmProjects/projectFlipkart/TestData/FlipkartDatabase.xlsx")
sh1=wb['LoginData']
mr_sh1=sh1.max_row   #maximum row in sheet
mc_sh1=sh1.max_column #max column in sheet
sh2=wb['Address']
mc_sh2=sh2.max_column
#print(type(sh1))
class xlData:
      user=sh1.cell(2,1).value
      password=sh1.cell(2,2).value

      F_name=sh2.cell(2,1).value
      L_name=sh2.cell(2,2).value
      name=sh2.cell(2,3).value
      mobile=sh2.cell(2,4).value
      pin=sh2.cell(2,5).value
      locality=sh2.cell(2,6).value
      address=sh2.cell(2,7).value










